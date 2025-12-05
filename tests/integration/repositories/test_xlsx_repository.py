# -*- coding: utf-8 -*-

# pylint: disable=import-error
# pylint: disable=missing-function-docstring

# Standard Library Imports
import pathlib
from typing import Any
from typing import Callable
from typing import Dict
from typing import Hashable
from typing import List
from typing import Optional

# Third-Party Imports
from openpyxl import load_workbook
import pytest

# Local Imports
from dodecahedron.models import AbstractModel
from dodecahedron.repositories import AbstractXlsxFileRepository
from dodecahedron.mappers import ClassMapper
from dodecahedron.testing import FakeRepository
from dodecahedron.wrappers import AbstractFileWrapper
from dodecahedron.wrappers import XlsxFileWrapper


class ExampleModel(AbstractModel):
    """Example model for testing."""

    def __init__(self, ref: str, value: str) -> None:
        self._reference = ref
        self._value = value

    @property
    def reference(self) -> str:
        """Reference."""
        return self._reference

    @property
    def value(self) -> str:
        """Value."""
        return self._value

    def __eq__(self, other: object) -> bool:
        return (
            other.reference == self.reference
            if isinstance(other, ExampleModel)
            else False
        )

    def __hash__(self) -> int:
        return hash(self._reference)

    def __repr__(self) -> str:
        result = f"<ExampleModel {self._reference}>"
        return result


class ExampleFileRepository(AbstractXlsxFileRepository, FakeRepository):
    """Example file repository for testing."""

    def __init__(
        self,
        wrapper: AbstractFileWrapper,
        /,
        mapper: Optional["ClassMapper[ExampleModel]"] = None,
        objects: Optional[List[ExampleModel]] = None,
    ) -> None:
        super().__init__(wrapper, mapper=mapper, objects=objects)

    def load(self) -> None:
        """Load objects from file."""
        records = self._read_records()
        objs = [self.mapper.from_dict(record) for record in records]  # type: ignore
        self._objects = set(objs)

    def save(self) -> None:
        """Save objects to file."""
        objects = sorted(
            list(self._objects),
            key=lambda v: getattr(v, "reference"),
        )
        records = [self.mapper.to_dict(obj) for obj in objects]  # type: ignore
        self._write_records(records)


@pytest.mark.parametrize("name", ["csv_file_wrapper", "txt_file_wrapper"])
def test_raises_error_when_not_an_xlsx_file(
    name: str, request: pytest.FixtureRequest
) -> None:
    with pytest.raises(TypeError, match="expected type 'XlsxFileWrapper'"):
        wrapper: AbstractFileWrapper = request.getfixturevalue(name)
        ExampleFileRepository(wrapper)


def test_loads_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [["id", "value"], ["1", "TEST"]]
    filepath = make_xlsx_file("test.xlsx", rows)
    wrapper = XlsxFileWrapper(filepath.resolve())
    schema: Dict[Hashable, Any] = {"id": "_reference", "value": "_value"}
    mapper = ClassMapper(ExampleModel, schema)
    repo = ExampleFileRepository(wrapper, mapper=mapper)
    repo.load()
    results = list(repo.objects)

    expected = [ExampleModel("1", "TEST")]
    assert results == expected


def test_saves_xlsx_file(tempdir: str) -> None:
    temppath = pathlib.Path(tempdir)
    filepath = temppath / "test.xlsx"
    wrapper = XlsxFileWrapper(filepath.resolve(), fieldnames=["id", "value"])
    schema: Dict[Hashable, Any] = {"id": "_reference", "value": "_value"}
    mapper = ClassMapper(ExampleModel, schema)
    repo = ExampleFileRepository(wrapper, mapper=mapper)

    repo.add(ExampleModel("1", "TEST"))
    repo.save()
    repo.commit()

    expected = temppath / "test.xlsx"
    assert expected.exists()


def test_adds_row_to_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [["id", "value"], ["1", "TEST"]]
    filepath = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(filepath.resolve(), fieldnames=["id", "value"])
    schema: Dict[Hashable, Any] = {"id": "_reference", "value": "_value"}
    mapper = ClassMapper(ExampleModel, schema)
    repo = ExampleFileRepository(wrapper, mapper=mapper)
    repo.load()

    repo.add(ExampleModel("2", "SUCCESS"))
    repo.save()
    repo.commit()

    workbook = load_workbook(filepath, data_only=True, read_only=True)
    worksheet = workbook.active
    result = worksheet["B3"].value  # type: ignore
    assert result == "SUCCESS"


# def test_adding_rows_is_idempotent(
#     make_xlsx_file: Callable[..., pathlib.Path]
# ) -> None:
#     rows = [["id", "value"], ["1", "TEST"]]
#     filepath = make_xlsx_file("test.xlsx", rows)

#     repo = AbstractXlsxBasedRepository(filepath)
#     obj = {"id": "2", "value": "SUCCESS"}
#     repo.add(obj)
#     repo.commit()

#     repo.add(obj)
#     repo.commit()

#     workbook = load_workbook(filepath, data_only=True, read_only=True)
#     worksheet = workbook.active
#     result = worksheet["B3"].value
#     assert result == "SUCCESS"
