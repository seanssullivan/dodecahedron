# -*- coding: utf-8 -*-

# pylint: disable=import-error
# pylint: disable=missing-function-docstring

# Standard Library Imports
import pathlib
from typing import Callable
from typing import Optional
from typing import Sequence

# Third-Party Imports
from openpyxl import Workbook
from openpyxl import load_workbook
import pytest

# Local Imports
from dodecahedron.wrappers.xlsx_file_wrappers import XlsxFileWrapper
from dodecahedron.wrappers.xlsx_file_wrappers import XlsxIOWrapper


def test_raises_error_when_filepath_argument_is_not_path() -> None:
    with pytest.raises(TypeError, match="expected type 'PathLike'"):
        XlsxFileWrapper("failure.txt")


def test_raises_error_when_filepath_argument_is_a_directory(
    make_txt_file: Callable[..., pathlib.Path],
) -> None:
    with pytest.raises(IsADirectoryError):
        path = make_txt_file("test.txt", "")
        XlsxFileWrapper(path.parent.resolve())


def test_raises_error_when_does_not_have_xlsx_extension(tempdir: str) -> None:
    path = pathlib.Path(tempdir) / "text.txt"
    with pytest.raises(ValueError, match="not a '.xlsx' file"):
        XlsxFileWrapper(path.resolve())


def test_opening_file_returns_xlsx_io_wrapper(
    make_xlsx_file: Callable[[str, list], pathlib.Path],
) -> None:
    path = make_xlsx_file("test.xlsx", [])
    wrapper = XlsxFileWrapper(path.resolve())
    result = wrapper.open()
    assert isinstance(result, XlsxIOWrapper)


def test_returned_io_wrapper_is_not_closed(
    make_xlsx_file: Callable[[str, list], pathlib.Path],
) -> None:
    path = make_xlsx_file("test.xlsx", [])
    wrapper = XlsxFileWrapper(path.resolve())
    result = wrapper.open()
    assert not result.closed


def test_can_be_passed_to_builtin_open_function(
    make_xlsx_file: Callable[[str, list], pathlib.Path],
) -> None:
    path = make_xlsx_file("test.xlsx", [])
    wrapper = XlsxFileWrapper(path.resolve())
    open(wrapper).close()


def test_loads_workbook_from_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open() as file:
        result = file.load_workbook()

    assert isinstance(result, Workbook)


def test_can_read_record_from_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open() as file:
        result = file.read_record()

    expected = {"id": "1", "value": "One"}
    assert result == expected


def test_reading_record_sets_fieldnames(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open() as file:
        file.read_record()

        expected = ["id", "value"]
        assert file.fieldnames == expected


def test_can_read_records_from_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open() as file:
        results = file.read_records()

    expected = [
        {"id": "1", "value": "One"},
        {"id": "2", "value": "Two"},
        {"id": "3", "value": "Three"},
    ]
    assert results == expected


def test_reading_records_sets_fieldnames(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open() as file:
        file.read_records()

        expected = ["id", "value"]
        assert file.fieldnames == expected


def test_can_read_row_from_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve(), fieldnames=["id", "value"])
    with wrapper.open() as file:
        result = file.read_row()

    expected = ["1", "One"]
    assert result == expected


@pytest.mark.parametrize("fieldnames", [None, ["id", "value"]])
def test_reading_row_does_not_return_fieldnames(
    make_xlsx_file: Callable[..., pathlib.Path],
    fieldnames: Optional[Sequence],
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve(), fieldnames=fieldnames)
    with wrapper.open() as file:
        result = file.read_row()

    expected = ["1", "One"]
    assert result == expected


def test_reading_row_sets_fieldnames(
    make_xlsx_file: Callable[..., pathlib.Path]
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open() as file:
        file.read_row()

        expected = ["id", "value"]
        assert file.fieldnames == expected


def test_can_read_rows_from_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve(), fieldnames=["id", "value"])
    with wrapper.open() as file:
        results = file.read_rows()

    expected = [["1", "One"], ["2", "Two"], ["3", "Three"]]
    assert results == expected


@pytest.mark.parametrize("fieldnames", [None, ["id", "value"]])
def test_reading_rows_does_not_return_headers(
    make_xlsx_file: Callable[..., pathlib.Path],
    fieldnames: Optional[Sequence],
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve(), fieldnames=fieldnames)
    with wrapper.open() as file:
        results = file.read_rows()

    expected = [["1", "One"], ["2", "Two"], ["3", "Three"]]
    assert results == expected


def test_reading_rows_sets_headers(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    rows = [
        ["id", "value"],
        ["1", "One"],
        ["2", "Two"],
        ["3", "Three"],
    ]
    path = make_xlsx_file("test.xlsx", rows)

    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open() as file:
        file.read_rows()

        expected = ["id", "value"]
        assert file.fieldnames == expected


def test_can_write_header_to_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    path = make_xlsx_file("test.xlsx")
    wrapper = XlsxFileWrapper(path.resolve(), fieldnames=["id", "value"])
    with wrapper.open("wb") as file:
        file.write_header()

    expected = [("id", "value")]
    worksheet = load_workbook(path).active
    results = [row for row in worksheet.iter_rows(values_only=True)]
    assert results == expected


def test_can_write_record_to_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    path = make_xlsx_file("test.xlsx")
    wrapper = XlsxFileWrapper(path.resolve(), fieldnames=["id", "value"])
    with wrapper.open("w") as file:
        file.write_header()
        file.write_record({"id": "1", "value": "One"})

    expected = [("id", "value"), ("1", "One")]
    worksheet = load_workbook(path).active
    results = [row for row in worksheet.iter_rows(values_only=True)]
    assert results == expected


def test_can_write_records_to_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    path = make_xlsx_file("test.xlsx")
    wrapper = XlsxFileWrapper(path.resolve(), fieldnames=["id", "value"])
    with wrapper.open("w") as file:
        file.write_header()
        file.write_records(
            [
                {"id": "1", "value": "One"},
                {"id": "2", "value": "Two"},
                {"id": "3", "value": "Three"},
            ]
        )

    expected = [("id", "value"), ("1", "One"), ("2", "Two"), ("3", "Three")]
    worksheet = load_workbook(path).active
    results = [row for row in worksheet.iter_rows(values_only=True)]
    assert results == expected


def test_can_write_row_to_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    path = make_xlsx_file("test.xlsx")
    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open("w") as file:
        file.write_row(["1", "One"])

    expected = [("1", "One")]
    worksheet = load_workbook(path).active
    results = [row for row in worksheet.iter_rows(values_only=True)]
    assert results == expected


def test_can_write_rows_to_csv_file(
    make_xlsx_file: Callable[..., pathlib.Path],
) -> None:
    path = make_xlsx_file("test.xlsx")
    wrapper = XlsxFileWrapper(path.resolve())
    with wrapper.open("w") as file:
        file.write_rows([["1", "One"], ["2", "Two"], ["3", "Three"]])

    expected = [("1", "One"), ("2", "Two"), ("3", "Three")]
    worksheet = load_workbook(path).active
    results = [row for row in worksheet.iter_rows(values_only=True)]
    assert results == expected
