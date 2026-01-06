# -*- coding: utf-8 -*-
"""Xlsx File Wrappers."""

# Standard Library Imports
from __future__ import annotations
import os
from typing import Any
from typing import Dict
from typing import IO
from typing import Iterable
from typing import Iterator
from typing import List
from typing import Optional
from typing import Sequence
from typing import Tuple
from typing import TypeVar
from typing import Union

# Third-Party Imports
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Local Imports
from .abstract_file_wrappers import AbstractDirectoryWrapper
from .abstract_file_wrappers import AbstractFileSystemWrapper
from .abstract_file_wrappers import AbstractFileWrapper
from .abstract_file_wrappers import AbstractIOWrapper
from ..utils import converters
from .. import helpers
from .. import settings
from .. import utils

__all__ = [
    "XlsxDirectoryWrapper",
    "XlsxFileWrapper",
    "XlsxIOWrapper",
]


# Custom type
T = TypeVar("T")


class AbstractXlsxWrapper(AbstractFileSystemWrapper):
    """Represents an abstract wrapper class for `.xlsx` files."""

    @property
    def fieldnames(self) -> Sequence[Any]:
        """Fieldnames."""
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value: Any) -> None:
        helpers.raise_for_instance(value, Sequence)
        self._fieldnames: Sequence[str] = value

    def _init_xlsx_io_wrapper(
        self, __file: IO[Any], /, mode: str
    ) -> XlsxIOWrapper:
        """Initialize I/O wrapper for `.xlsx` file.

        Args:
            __file: File-like object.
            mode: File mode.

        Returns:
            I/O wrapper instance.

        """
        result = XlsxIOWrapper(__file, mode=mode)
        setattr(result, "_context", self)
        return result


class XlsxDirectoryWrapper(AbstractXlsxWrapper, AbstractDirectoryWrapper):
    """Implements a wrapper for `.xlsx` files in a directory.

    Args:
        directory: Directory from which to load `.xlsx` file(s).
        fieldnames (optional): Fieldnames. Default ``None``.
        read_only (optional): Whether file is read only. Default ``False``.

    """

    def __init__(
        self,
        directory: "os.PathLike[Any]",
        *,
        fieldnames: Optional[Sequence[str]] = None,
        read_only: bool = False,
    ) -> None:
        super().__init__(
            directory,
            extension=settings.XLSX_EXTENSION,
            read_only=read_only,
        )
        self.fieldnames = fieldnames or []

    def open(self, filename: str, /, mode: str = "rb") -> XlsxIOWrapper:
        """Open a `.xlsx` file and return a file object.

        Args:
            filename: Filename.
            mode (optional): File mode. Default ``rb``.

        Returns:
            File object.

        Raises:
            TypeError: when `encoding` is not type ``str``.

        """

        file = super().open(filename, mode=converters.to_bytes_file_mode(mode))
        result = self._init_xlsx_io_wrapper(file, mode=mode)
        return result


class XlsxFileWrapper(AbstractXlsxWrapper, AbstractFileWrapper):
    """Implements a wrapper for `.xlsx` files.

    Args:
        filepath: Path to `.xlsx` file.
        fieldnames (optional): Fieldnames. Default ``None``.
        read_only (optional): Whether file is read only. Default ``False``.

    Raises:
        ValueError: when `filepath` is not a `.xlsx` file.

    """

    def __init__(
        self,
        filepath: "os.PathLike[Any]",
        *,
        fieldnames: Optional[Sequence[str]] = None,
        read_only: bool = False,
    ) -> None:
        super().__init__(filepath, read_only=read_only)
        utils.raise_for_extension(filepath, settings.XLSX_EXTENSION)

        self.fieldnames = fieldnames or []

    def open(self, mode: str = "rb") -> XlsxIOWrapper:
        """Open the `.xlsx` file and return a file object.

        Args:
            mode (optional): File mode. Default ``rb``.

        Returns:
            File object.

        """
        file = super().open(converters.to_bytes_file_mode(mode))
        result = self._init_xlsx_io_wrapper(file, mode=mode)
        return result


class OpenPyXLMixin:
    """Implements an OpenPyXL mixin."""

    _file: IO[Any]
    _mode: str

    def __exit__(self, *_: Any) -> None:
        read_only: bool = getattr(self, "read_only")
        if "w" in self._mode and not read_only:
            self.save()

    def get_worksheet(self) -> Worksheet:
        """Get worksheet.

        Returns:
            Worksheet.

        """
        workbook = self.get_workbook()
        sheetname = getattr(self, "sheetname", None)
        result = workbook[sheetname] if sheetname else workbook.active
        return result  # type: ignore

    def get_workbook(self) -> Workbook:
        """Get workbook.

        Returns:
            Workbook.

        """
        result = getattr(self, "_workbook", None)
        if not result and "r" in self._mode:
            result = self.load_workbook()

        if not result and "w" in self._mode:
            result = self.make_workbook()

        return result  # type: ignore

    def load_workbook(self) -> Workbook:
        """Load workbook.

        Returns:
            Workbook.

        """
        read_only: bool = getattr(self, "read_only")
        result = load_workbook(self._file, read_only=read_only)
        setattr(self, "_workbook", result)
        return result

    def make_workbook(self) -> Workbook:
        """Make workbook.

        Returns:
            Workbook.

        """
        result = Workbook()
        setattr(self, "_workbook", result)
        return result

    def save(self) -> None:
        """Save workbook."""
        workbook = self.get_workbook()
        workbook.save(self._file)


class XlsxIOWrapper(OpenPyXLMixin, AbstractIOWrapper):
    """Implements a I/O wrapper for `.xlsx` files."""

    def __init__(self, __file: IO[Any], /, mode: str) -> None:
        self._file = __file
        self._mode = mode
        self._context: Optional[AbstractXlsxWrapper] = None

    @property
    def file(self) -> IO[Any]:
        """File."""
        return self._file

    @property
    def closed(self) -> bool:
        """Whether file is closed."""
        return self._file.closed

    @property
    def context(self) -> Union[AbstractXlsxWrapper, XlsxIOWrapper]:
        """Context."""
        return self._context or self

    @property
    def fieldnames(self) -> Sequence[str]:
        """Fieldnames."""
        result = getattr(self.context, "_fieldnames", [])
        return result

    @fieldnames.setter
    def fieldnames(self, value: Any) -> None:
        helpers.raise_for_instance(value, Sequence)
        setattr(self.context, "_fieldnames", value)

    @property
    def sheetname(self) -> Optional[str]:
        """Sheetname."""
        result = getattr(self, "_sheetname", None)
        return result

    @sheetname.setter
    def sheetname(self, value: Any) -> None:
        if value is not None and not isinstance(value, str):
            message = f"expected type 'str', got {type(value)} instead"
            raise TypeError(message)

        setattr(self, "_sheetname", value)

    @property
    def read_only(self) -> bool:
        """Whether read only."""
        return getattr(self._context, "read_only")

    def __enter__(self) -> XlsxIOWrapper:
        return self

    def __exit__(self, *_: Any) -> None:
        super().__exit__(*_)
        self.close()
        return

    def close(self) -> None:
        """Close `.xlsx` file."""
        self._file.close()

    def read(self, n: int = -1, /) -> bytes:
        """Read content of `.xlsx` file.

        Returns:
            Content.

        """
        result = self._file.read(n)
        return result

    def read_record(self) -> Dict[str, Any]:
        """Read record from `.xlsx` file.

        Returns:
            Record.

        """
        reader = self._get_record_reader()
        result: Dict[str, Any] = reader.read_record()

        if not self.fieldnames:
            self.fieldnames = reader.fieldnames

        return result

    def read_records(self) -> List[Dict[str, Any]]:
        """Read records from `.xlsx` file.

        Returns:
            Records.

        """
        reader = self._get_record_reader()
        results = reader.read_records()

        if not self.fieldnames:
            self.fieldnames = reader.fieldnames

        return results

    def _get_record_reader(self) -> _OpenPyXLRecordReader:
        """Get `.xlsx` record reader.

        Returns:
            Reader.

        """
        if not hasattr(self, "_record_reader"):
            self._start_record_reader()

        result: _OpenPyXLRecordReader = getattr(self, "_record_reader")
        return result

    def _start_record_reader(self) -> None:
        """Start `.xlsx` record reader."""
        reader = _OpenPyXLRecordReader(self)
        setattr(self, "_record_reader", reader)

    def read_row(self) -> List[str]:
        """Read row from `.xlsx` file.

        Returns:
            Row.

        """
        reader = self._get_row_reader()
        result = reader.read_row()

        if not self.fieldnames:
            self.fieldnames = reader.fieldnames

        return result

    def read_rows(self) -> List[List[str]]:
        """Read rows from `.xlsx` file.

        Returns:
            Rows.

        """
        reader = self._get_row_reader()
        results = reader.read_rows()

        if not self.fieldnames:
            self.fieldnames = reader.fieldnames

        return results

    def _get_row_reader(self) -> _OpenPyXLRowReader:
        """Get `.xlsx` row reader.

        Returns:
            Reader.

        """
        if not hasattr(self, "_row_reader"):
            self._start_row_reader()

        result: _OpenPyXLRowReader = getattr(self, "_row_reader")
        return result

    def _start_row_reader(self) -> None:
        """Start `.xlsx` row reader."""
        reader = _OpenPyXLRowReader(self)
        setattr(self, "_row_reader", reader)

    def write(self, buffer: memoryview, /) -> int:  # type: ignore
        """Write buffer to `.xlsx` file.

        Args:
            buffer: Buffer.

        """
        result = self._file.write(buffer)
        return result

    def write_header(self) -> None:
        """Write header."""
        writer = self._get_record_writer()
        writer.write_header()

    def write_record(self, record: Dict[str, Any], /) -> None:
        """Write record to `.xlsx` file.

        Args:
            record: Record to write.

        """
        writer = self._get_record_writer()
        writer.write_record(record)

    def write_records(self, records: Iterable[Dict[str, Any]], /) -> None:
        """Write records to `.xlsx` file.

        Args:
            records: Records to write.

        """
        writer = self._get_record_writer()
        writer.write_records(records)

    def _get_record_writer(self) -> _OpenPyXLRecordWriter:
        """Get record writer for `.xlsx` file.

        Returns:
            Writer.

        """
        if not hasattr(self, "_record_writer"):
            self._start_record_writer()

        result: _OpenPyXLRecordWriter = getattr(self, "_record_writer")
        return result

    def _start_record_writer(self) -> None:
        """Start record writer for `.xlsx` file."""
        writer = _OpenPyXLRecordWriter(self)
        setattr(self, "_record_writer", writer)

    def write_row(self, row: Iterable[Any], /) -> None:
        """Write row to `.xlsx` file.

        Args:
            row: Row to write.

        """
        writer = self._get_row_writer()
        writer.write_row(row)

    def write_rows(self, rows: Iterable[Iterable[Any]], /) -> None:
        """Write rows to `.xlsx` file.

        Args:
            rows: Rows to write.

        """
        writer = self._get_row_writer()
        writer.write_rows(rows)

    def _get_row_writer(self) -> _OpenPyXLRowWriter:
        """Get row `.xlsx` writer.

        Returns:
            Writer.

        """
        if not hasattr(self, "_writer"):
            self._start_row_writer()

        result: _OpenPyXLRowWriter = getattr(self, "_writer")
        return result

    def _start_row_writer(self) -> None:
        """Start standard `.xlsx` writer."""
        writer = _OpenPyXLRowWriter(self)
        setattr(self, "_writer", writer)


class _OpenPyXLRecordReader(Iterator[Any]):
    """Implements an OpenPyXL record reader for `.xlsx` files."""

    def __init__(self, __wrapper: "XlsxIOWrapper") -> None:
        self._worksheet = __wrapper.get_worksheet()
        self._reader = self._worksheet.iter_rows(values_only=True)
        self._fieldnames = __wrapper.fieldnames
        self._row_num = 0

    @property
    def fieldnames(self) -> Sequence[str]:
        """Fieldnames."""
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, fieldnames: Sequence[str]) -> None:
        self._fieldnames = fieldnames

    @property
    def row_num(self) -> int:
        """Row number."""
        return self._row_num

    def __next__(self) -> Dict[str, Any]:
        result = self.read_record()
        return result

    def read_record(self) -> Dict[str, Any]:
        """Read record from `.csv` file.

        Returns
            Record.

        """
        if self.row_num == 0 and self.fieldnames:
            result = self._read_first_record()
            return result

        if self.row_num == 0 and not self.fieldnames:
            self._read_header()

        result = self._read_next_record()
        return result

    def read_records(self) -> List[Dict[str, Any]]:
        """Read rows from `.xlsx` file.

        Returns:
            Rows.

        """
        if self.row_num == 0 and self.fieldnames:
            first_row = self._read_first_record()
            remainder = [self._make_record(row) for row in self._reader]
            self._row_num += len(remainder)
            results = [first_row, *remainder]
            return results

        if self.row_num == 0 and not self.fieldnames:
            self._read_header()

        results = [self._make_record(row) for row in self._reader]
        self._row_num += len(results)
        return results

    def _read_header(self) -> None:
        """Read header.

        Raises:
            RuntimeError: when `fieldnames` is not empty.
            RuntimeError: when not on first line of file.

        """
        if self.fieldnames:  # sanity check
            message = "'fieldnames' is not empty"
            raise RuntimeError(message)

        if self.row_num != 0:
            message = "reader is not on first line of file"
            raise RuntimeError(message)

        self.fieldnames = self._read_next_row()

    def _read_next_row(self) -> List[str]:
        """Read next row.

        Returns:
            Row.

        """
        row = next(self._reader)
        result = self._make_row(row)
        self._row_num += 1
        return result

    def _make_row(self, __row: Tuple[Any, ...], /) -> List[Any]:
        """Make row.

        Returns:
            Row.

        """
        result = list(__row)
        return result

    def _read_first_record(self) -> Dict[str, Any]:
        """Read first row.

        Args:
            reader: Reader.

        Raises:
            RuntimeError: when `reader` is not on first line of file.

        """
        if self.row_num != 0:
            message = "reader is not on first line of file"
            raise RuntimeError(message)

        first_record = self._read_next_record()
        if not self._is_header(list(first_record.values())):
            return first_record

        result = self._read_next_record()
        return result

    def _read_next_record(self) -> Dict[str, str]:
        """Read next record.

        Returns:
            Row.

        """
        row = next(self._reader)
        result = self._make_record(row)
        self._row_num += 1
        return result

    def _make_record(self, __row: Tuple[Any, ...], /) -> Dict[str, Any]:
        """Make record from row.

        Args:
            row: Row.

        Returns:
            Record.

        """
        result = {key: value for key, value in zip(self.fieldnames, __row)}
        return result

    def _is_header(self, row: Sequence[str]) -> bool:
        """Check whether row is header.

        Args:
            row: Row.

        Returns:
            Whether row is header.

        Raises:
            RuntimeError: when `fieldnames` is empty.

        """
        if not self.fieldnames:  # sanity check
            message = "'fieldnames' is empty"
            raise RuntimeError(message)

        result = not set(self.fieldnames).difference(row)
        return result


class _OpenPyXLRecordWriter:
    """Implements an OpenPyXL record writer for `.xlsx` files."""

    def __init__(self, __wrapper: "XlsxIOWrapper") -> None:
        self._worksheet = __wrapper.get_worksheet()
        self._fieldnames = __wrapper.fieldnames
        self._row_num = 0

    @property
    def fieldnames(self) -> Sequence[str]:
        """Fieldnames."""
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, fieldnames: Sequence[str]) -> None:
        self._fieldnames = fieldnames

    @property
    def row_num(self) -> int:
        """Row number."""
        return self._row_num

    def write_header(self) -> None:
        """Write header."""
        row = self._row_num + 1  # OpenPyXl rows are not zero-indexed
        for col, fieldname in enumerate(self.fieldnames, start=1):
            self._worksheet.cell(row, col, value=fieldname)

        self._row_num += 1

    def write_record(self, __record: Dict[str, Any], /) -> None:
        """Write record to `.xlsx` file.

        Args:
            __record: Record to write.

        """
        row = self._row_num + 1  # OpenPyXl rows are not zero-indexed
        for col, key in enumerate(self.fieldnames, start=1):
            self._worksheet.cell(row, col, value=__record[key])

        self._row_num += 1

    def write_records(self, __records: Iterable[Dict[str, Any]], /) -> None:
        """Write records to `.csv` file.

        Args:
            __records: Records to write.

        """
        start = self._row_num + 1  # OpenPyXl rows are not zero-indexed
        for row, record in enumerate(__records, start=start):
            for col, key in enumerate(self.fieldnames, start=1):
                self._worksheet.cell(row, col, value=record[key])

        self._row_num += len(list(__records))


class _OpenPyXLRowReader(Iterator[Any]):
    """Implements an OpenPyXL row reader for `.xlsx` files."""

    def __init__(self, __wrapper: "XlsxIOWrapper") -> None:
        self._worksheet = __wrapper.get_worksheet()
        self._reader = self._worksheet.iter_rows(values_only=True)
        self._fieldnames = __wrapper.fieldnames
        self._row_num = 0

    @property
    def fieldnames(self) -> Sequence[str]:
        """Fieldnames."""
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, fieldnames: Sequence[str]) -> None:
        self._fieldnames = fieldnames

    @property
    def row_num(self) -> int:
        """Row number."""
        return self._row_num

    def __next__(self) -> List[Any]:
        result = self.read_row()
        return result

    def read_row(self) -> List[str]:
        """Read row from `.csv` file.

        Returns
            Row.

        """
        if self.row_num == 0 and self.fieldnames:
            result = self._read_first_row()
            return result

        if self.row_num == 0 and not self.fieldnames:
            self._read_header()

        result = self._read_next_row()
        return result

    def read_rows(self) -> List[List[str]]:
        """Read rows from `.csv` file.

        Returns:
            Rows.

        """
        if self.row_num == 0 and self.fieldnames:
            first_row = self._read_first_row()
            remainder = [self._make_row(row) for row in self._reader]
            self._row_num += len(remainder)
            results: List[List[Any]] = [first_row, *remainder]
            return results

        if self.row_num == 0 and not self.fieldnames:
            self._read_header()

        results = [self._make_row(row) for row in self._reader]
        self._row_num += len(results)
        return results

    def _read_header(self) -> None:
        """Read header.

        Raises:
            RuntimeError: when `fieldnames` is not empty.
            RuntimeError: when not on first line of file.

        """
        if self.fieldnames:  # sanity check
            message = "'fieldnames' is not empty"
            raise RuntimeError(message)

        if self.row_num != 0:
            message = "reader is not on first line of file"
            raise RuntimeError(message)

        self.fieldnames = self._read_next_row()

    def _read_first_row(self) -> List[str]:
        """Read first row.

        Args:
            reader: Reader.

        Raises:
            RuntimeError: when `reader` is not on first line of file.

        """
        if self.row_num != 0:
            message = "reader is not on first line of file"
            raise RuntimeError(message)

        first_row = self._read_next_row()
        if not self._is_header(first_row):
            return first_row

        result = self._read_next_row()
        return result

    def _read_next_row(self) -> List[str]:
        """Reqad next row.

        Returns:
            Row.

        """
        row = next(self._reader)
        result = self._make_row(row)
        self._row_num += 1
        return result

    def _make_row(self, __row: Tuple[Any, ...], /) -> List[Any]:
        """Make row.

        Returns:
            Row.

        """
        result = list(__row)
        return result

    def _is_header(self, row: Sequence[str]) -> bool:
        """Check whether row is header.

        Args:
            row: Row.

        Returns:
            Whether row is header.

        Raises:
            RuntimeError: when `fieldnames` is empty.

        """
        if not self.fieldnames:  # sanity check
            message = "'fieldnames' is empty"
            raise RuntimeError(message)

        result = not set(self.fieldnames).difference(row)
        return result


class _OpenPyXLRowWriter:
    """Implements an OpenPyXL row writer for `.xlsx` files."""

    def __init__(self, __wrapper: "XlsxIOWrapper") -> None:
        self._worksheet = __wrapper.get_worksheet()
        self._fieldnames = __wrapper.fieldnames
        self._row_num = 0

    @property
    def fieldnames(self) -> Sequence[str]:
        """Fieldnames."""
        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, fieldnames: Sequence[str]) -> None:
        self._fieldnames = fieldnames

    @property
    def row_num(self) -> int:
        """Row number."""
        return self._row_num

    def write_header(self) -> None:
        """Write header."""
        self.write_row(self.fieldnames)

    def write_row(self, __row: Iterable[Any], /) -> None:
        """Write row to `.xlsx` file.

        Args:
            row: Row to write.

        """
        row = self._row_num + 1  # OpenPyXl rows are not zero-indexed
        for col, fieldname in enumerate(__row, start=1):
            self._worksheet.cell(row, col, value=fieldname)

        self._row_num += 1

    def write_rows(self, __rows: Iterable[Iterable[Any]], /) -> None:
        """Write rows to `.xlsx` file.

        Args:
            __rows: Rows to write.

        """
        start = self._row_num + 1  # OpenPyXl rows are not zero-indexed
        for row, data in enumerate(__rows, start=start):
            for col, value in enumerate(data, start=1):
                self._worksheet.cell(row, col, value=value)

        self._row_num += len(list(__rows))
